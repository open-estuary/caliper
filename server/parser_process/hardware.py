#!/usr/bin/python
import os
import yaml
import openpyxl
import glob
from caliper.client.shared import caliper_path
path = caliper_path.HW_DATA_DIR_INPUT
template = os.path.join(caliper_path.TEMPLATE_DATA_DIR,"Hw_Template.xlsx")
output = os.path.join(caliper_path.HW_DATA_DIR_OUTPUT,'Platform_Configuration.xlsx')
dic = {
		'CPU':{
			'description' : ['Model_Name','Sockets','Cores_Per_Socket','Threads_per_Core','CPU_Cores','Numa_Node','BogoMIPS'],
			'col' : 2
				},
		'DISK':{
            'description' : ['RAID_Card','RAID_Vendor','sda_Model','sda_Size','sda_Serial','sda_Vendor','sdb_Model','sdb_Size','sdb_Serial','sdb_Vendor'],
			'col' : 3
				},
        'NETWORK':{
			'description':['PCI_Gigabit_Card'],
			'col' : 2
				},
        'OS':{
			'description' : ['Release','Codename','Description','GCC_Version','LD_Version'],
			'col' : 2
				},
        'KERNEL':{
			'description' : ['Version'],
			'col' : 2
				},
		'MEMORY':{
			'description' : ['L1_D-Cache_Size','L1_D-Cache_Associativity','L1_D-Cache_Operational_Mode','L1_I-Cache_Size','L1_I-Cache_Associativity','L1_I-Cache_Operational_Mode','L2_Cache_Size','L2_Cache_Associativity','L2_Cache_Operational_Mode','L3_Cache_Size','L3_Cache_Associativity','L3_Cache_Operational_Mode','Main_Memory_Type','Main_Memory_Formfactor','Main_Memory_Max_Speed','Main_Memory_Current_Speed','Main_Memory_Manufacturer','Main_Memory_Part_number','Main_Memory_Size'],
			'col' : 3
				}	
			}


def populate_excel():
    wb = openpyxl.load_workbook(template)
    col_add = 0
    list_index = 0
    if not os.path.exists(caliper_path.HW_DATA_DIR_OUTPUT):
        os.makedirs(caliper_path.HW_DATA_DIR_OUTPUT)
    for files in glob.glob(os.path.join(path,"*yaml")):
        for sheet_name,desc in dic.iteritems():
            desc_list = dic[sheet_name]['description']	
            col = dic[sheet_name]['col']
            sheet = wb.get_sheet_by_name(sheet_name)
            dic_hw = yaml.load(open(files))
            #for i in range(3,len(desc_list)+2):
            i = 3
            sheet.cell(row = 2,column = col+col_add).value = "_".join(files.split("/")[-1].split("_")[:-2])
            for list_index in range(0,len(desc_list)):
                sheet.cell(row = i,column = col+col_add).value = dic_hw['Hardware_Info'][sheet_name][desc_list[list_index]]
                i += 1
        col_add += 1
    wb.save(output)
    

if __name__ == "__main__":
    populate_excel()

