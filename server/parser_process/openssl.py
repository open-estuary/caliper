#!/usr/bin/python
import os
import re
import pdb
import openpyxl
import glob
import sys
import logging
from caliper.client.shared import caliper_path

template_dir = caliper_path.TEMPLATE_DATA_DIR
template = os.path.join(template_dir,"Openssl_Template.xlsx")

input_path = caliper_path.OPENSSL_DATA_DIR_INPUT

output_path = caliper_path.EXCEL_DATA_DIR_OUTPUT
output_excel = os.path.join(output_path,'openssl_TestCases.xlsx')

def openssl_populate(filename,column,dic,wb):
    for sheet,desc in dic.iteritems():
        sheet_name = sheet
        desc_list = dic[sheet]['description']
        table_index = dic[sheet]['table_index']
        reg_ex = dic[sheet]['reg_ex']
        sheet = wb.get_sheet_by_name(sheet_name)
        #populates the heading for the given column
        sheet.cell(row = 1,column = column).value = "_".join(filename.split("/")[-1].split("_")[:-1])
        with open(filename,'r') as fp:
            content = fp.read()
            contents = content.split("type")
            tables = re.split(r'\s{18}[so]',contents[1])
            r = 2
            row_index = 2
            #excel_sign_keys = sheet.col_values(1,2,45)
            col_list = []
            #takes all the keys of column1
            while sheet.cell(row = row_index, column = 1).value != None :
                col_list.append((sheet.cell(row = row_index, column = 1).value))
                row_index +=1
            for j in table_index:
                for i in range(1,len(tables[j].splitlines())):
                    if re.search(reg_ex,tables[j].splitlines()[i]):
                        keys = re.search(reg_ex,tables[j].splitlines()[i])
                        if keys.group(1).strip(" ") in col_list:
                            for row_no in range(0,len(desc_list)):
                                index = col_list.index(keys.group(1).strip(" "))
                                sheet.cell(row=index+r+row_no,column=column).value = keys.group(row_no+2)#this will print for all the values in
                                col_list[:] = (value for value in col_list  if value != keys.group(row_no+2))
def merge(wb,r,c):
    sheets = wb.get_sheet_names()
    for sheet_name in sheets:
        sheet = wb.get_sheet_by_name(sheet_name)
        row = r
        column = c
        while 1:
            data = sheet.cell(row = row, column = column).value
            prev_row = row
            if not data:
                break
            while sheet.cell(row = row, column = column).value == data:
                row +=1
            sheet.merge_cells(start_row=prev_row,start_column=column,end_row=row-1,end_column=column)
            currentCell = sheet.cell(row=prev_row, column=column).coordinate
            sheet[currentCell].alignment = openpyxl.styles.Alignment(horizontal='center')
            sheet[currentCell].alignment = openpyxl.styles.Alignment(vertical='center')
    wb.save(output_excel)

def populate_excel():
    dic = {

            'DigitalVerify':{
            'description':['verify','verify/s'],
            #'reg_ex':r'([0-9a-z()\s]+)\d+\.\d+s\s+(\d+\.\d+s)\s+\d+\.\d+\s+(\d+\.\d+)',
            'reg_ex': r'([0-9a-z()\s]+)\d+\.\d+s\s+(\d+\.\d+)s\s+\d+\.\d+\s+(\d+\.\d+)',
            'table_index':range(1,4)
            },

        'DigitalSign':{
            'description':['sign','sign/s'],
            #'reg_ex':r'([0-9a-z()\s]+)(\d+\.\d+s)\s+\d+\.\d+s\s+(\d+\.\d+)\s+\d+\.\d+',
            'reg_ex': r'([0-9a-z()\s]+)(\d+\.\d+)s\s+\d+\.\d+s\s+(\d+\.\d+)\s+\d+\.\d+',
            'table_index':range(1,4)
            },
        'ECDH':{
            'description':['operations','operations/s'],
            #'reg_ex':r'([0-9a-z()\s]+)(\d+\.\d+s)\s+(\d+\.\d+)',
            'reg_ex': r'([0-9a-z()\s]+)(\d+\.\d+)s\s+(\d+\.\d+)',
            'table_index':range(4,5)
        },
        'SymmetricCypher':{
            'description':['16 bytes','64 bytes','256 bytes','1024 bytes','8192 bytes'],
            #'reg_ex':r'([0-9a-z()\-\s]+)\s{2}(\d+\.\d+k*)\s{1}\s*(\d+\.\d+k*)\s{1}\s*(\d+\.\d+k*)\s{1}\s*(\d+\.\d+k*)\s{1}\s*(\d+\.\d+k*)',
            'reg_ex': r'([0-9a-z()\-\s]+)\s{2}(\d+\.\d+)k*\s{1}\s*(\d+\.\d+)k*\s{1}\s*(\d+\.\d+)k*\s{1}\s*(\d+\.\d+)k*\s{1}\s*(\d+\.\d+)k*',
            'table_index': range(0,1)
        },
        'HashAlgorithms':{
            'description':['16 bytes','64 bytes','256 bytes','1024 bytes','8192 bytes'],
            #'reg_ex':r'([0-9a-z()\-\s]+)\s{2}(\d+\.\d+k*)\s{2}\s*(\d+\.\d+k*)\s{2}\s*(\d+\.\d+k*)\s{2}\s*(\d+\.\d+k*)\s{2}\s*(\d+\.\d+k*)',
            'reg_ex': r'([0-9a-z()\-\s]+)\s{2}(\d+\.\d+)k*\s{2}\s*(\d+\.\d+)k*\s{2}\s*(\d+\.\d+)k*\s{2}\s*(\d+\.\d+)k*\s{2}\s*(\d+\.\d+)k*',
            'table_index': range(0, 1)
        }
        }
    #populating the value will take place from 3rd column the other two has descriptions
    column_start = 3
    wb = openpyxl.load_workbook(template)
    if not os.path.exists(output_path):
        os.makedirs(output_path)
    for files in glob.glob(os.path.join(input_path,"*log")):

        try:
            openssl_populate(files,column_start,dic,wb)
            merge(wb,2,1)
        except Exception as e:
            logging.info(e)
            raise
        else:
            column_start += 1
if __name__ == "__main__":
    populate_excel()
