#!/usr/bin/python
import os
import yaml
import openpyxl
import glob
#from caliper.client.shared import caliper_path
import numpy
import sys
from openpyxl.utils import (_get_column_letter)

col_start = 6
row_start = 3

keys = [ "Functional","Performance"]

functional_list = ["kernel","debug"]
perf_list = ["algorithm","application","cpu_sincore","cpu_multicore","storage","latency","memory","network"]
del_list = ["peripheral"]



def populate_excel_to_yaml(row,c,dic,scenario_col,testcase_col,key,wb,file_name):
    # adding exception block to handill the exception during the web report generation
    # when any one of Functional or Performance is missing in .yaml files
    try:
      for i in range(0,len(dic['results'][key].keys())):
        sheet_name = dic['results'][key].keys()[i]

        if sheet_name in del_list:
            continue
        sheet = wb.get_sheet_by_name(sheet_name)
        r = row
        head_row = row-1
        sheet.cell(row = head_row, column = c).value = file_name

        while(sheet.cell(row = r,column = scenario_col).value != None):

            try:
                if(dic['results'][key][sheet_name][sheet.cell(row = r,column = scenario_col).value][sheet.cell(row = r,column = testcase_col).value] >0):
                    sheet.cell(row = r,column = c).value = dic['results'][key][sheet_name][sheet.cell(row = r,column = scenario_col).value][sheet.cell(row = r,column = testcase_col).value]
                else:
                    sheet.cell(row=r, column=c).value = "Missing"
            except:
                sheet.cell(row = r,column = c).value = "Missing"
            r += 1
    except Exception as e:
        pass

def get_cov_file_list(input_cov):
    file_dir_list = []
    file_name_list = []

    files_list = []
    for x in os.walk(input_cov):
        file_dir_list.append(x[0])
        file_name_list.append(x[2])
    del file_dir_list[0]


    # save the count of number directories inside Input_Cov
    len_file_name_list = len(file_name_list)
    # Take the yaml file list from Input_Cov/1 directory, as this directory contains
    # yaml files. Because at least one iteration should have execute to generate report.

    min_file_count = 255
    index = 0

    for i in range(0, len_file_name_list):
        if len(file_name_list[i]):
            if min_file_count > len(file_name_list[i]):
                min_file_count = len(file_name_list[i])
                index = i

    for files in file_name_list[index]:
        files_temp_list = []
        i = 1
        for dir in file_dir_list:
            if len(file_name_list[i]):
                file_name = os.path.join(dir, files)
                files_temp_list.append(file_name)
            i += 1
        files_list.append(files_temp_list)

    return files_list

def row_slice(wb,sheet_name,row,col_start,col_end):
    data_list = []
    a = (_get_column_letter(col_start))
    b = (_get_column_letter(col_end))
    a += str(row)
    b += str(row)
    col_range = a+":"+b
    cell_list = tuple(sheet_name.iter_rows(col_range))

    for cell_item in cell_list:
        for item in cell_item:

            data_list.append(item.value)
    data_list = filter(data_list)
    return data_list

def mean_cov_populate(input_excel_path,col_end,file_name):
    col_s = col_start
    col_e = col_end -1

    for key in keys:
        excel_name = key + "_" + file_name + "-Tests.xlsx"
        output_excel = os.path.join(input_excel_path,excel_name)
        rb = openpyxl.load_workbook(output_excel)
        list_sheet = rb.get_sheet_names()


        for sheets in list_sheet:

            rb_sheet = rb.get_sheet_by_name(sheets)
            row_s = row_start
            rb_sheet.cell(row=row_s-1, column=col_end + 2).value = "Cov"
            rb_sheet.cell(row=row_s-1, column=col_end + 1).value = "Mean"
            while (rb_sheet.cell(row=row_s, column=col_s-2).value != None):


                data_list = row_slice(rb,rb_sheet,row_s,col_start,col_end)

                if data_list:
                    mean_val = numpy.mean(data_list)
                    rb_sheet.cell(row = row_s,column = col_end+1).value = mean_val
                    stdev_val = numpy.std(data_list)
		    # Update the covariance equation to express values in percentage
                    cov= (stdev_val / mean_val) * 100
                    rb_sheet.cell(row = row_s, column = col_end+ 2).value = cov

                    data_list = []
                else:
                    rb_sheet.cell(row=row_s, column=col_end + 2).value = "Missing"
                    rb_sheet.cell(row=row_s, column=col_end + 1).value = "Missing"
                row_s +=1


        rb.save(output_excel)

def filter(a):
    data_list=[]
    allowed_type =["<type 'long'>","<type 'float'>","<type 'int'>"]
    for item in a:
        if str(type(item))in allowed_type:
            data_list.append(item)
    return data_list


def sort_file_list(file_list):

    #for files in glob.glob(os.path.join(input_path, "*yaml")):
    #    file_list.append(files)

    files_new = []
    for i in range(0, len(file_list)):
        fp = open(file_list[i])
        data = yaml.load(fp)
        data = data['Configuration']
        if data['Machine_arch']:
            if data['Machine_arch'] == 'x86_64':
                files_new.append(file_list[i])
    files_new.sort()
    file_list = [i for i in file_list if not i in files_new]
    file_list.sort()
    files_new.sort()
    for i in files_new:
        file_list.append(i)
    return file_list

def populate_excel(ip_file_list,output_path,template_dir,cov_flag):
    scenario_column = 3
    testcase_column = 4
    col = col_start
    if not os.path.exists(output_path):
        os.makedirs(output_path)
    for key in keys:
        excel_name = key + "_Template.xlsx"
        if cov_flag:
            excel_name_save = key + "_" + ".".join(ip_file_list[0].split("/")[-1].split(".")[0:-1]) + "-Tests.xlsx"
        else:
            excel_name_save = key + "-Tests.xlsx"
        template = os.path.join(template_dir, excel_name)
        output_excel = os.path.join(output_path, excel_name_save)

        wb = openpyxl.load_workbook(template)
        file_list = sort_file_list(ip_file_list)
        for files in file_list:
            dic = yaml.load(open(files))
            file_name = ".".join(files.split("/")[-1].split(".")[0:-1])
            if cov_flag:
                file_name +=  "_" + files.split("/")[-2]
            populate_excel_to_yaml(row_start, col, dic, scenario_column, testcase_column, key, wb, file_name)
            col += 1
        col_end = col
        col = col_start
        wb.save(output_excel)
    return col_end-1

def get_COV_excel(input_excel_path,file_name,output_excel_path,col,template_dir,Iteration_len):
    for key in keys:
        input_excel_name =  key + "_" + file_name + "-Tests.xlsx"

        input_excel_name = os.path.join(input_excel_path,input_excel_name)
        output_excel_name = os.path.join(output_excel_path, key + "_" + "cov" + ".xlsx" )
        wb = openpyxl.load_workbook(input_excel_name)

        if col == col_start:
            excel_name = os.path.join(template_dir, key + "_Template.xlsx")
        else:
            excel_name = output_excel_name

        rb = openpyxl.load_workbook(excel_name)
        rows = row_start
        for sheets in wb.get_sheet_names():
            #rb = openpyxl.load_workbook(output_excel_path)
            ip_sheet = wb.get_sheet_by_name(sheets)
            op_sheet = rb.get_sheet_by_name(sheets)
            op_sheet.cell(row=row_start-1, column=col).value = file_name
            while ip_sheet.cell(row= rows, column = 7 + Iteration_len).value != None:
                op_sheet.cell(row = rows, column = col ).value = ip_sheet.cell(row= rows,column = 7 + Iteration_len).value
                rows += 1
            rows = row_start
        rb.save(output_excel_name)

if __name__ == "__main__" :
    input_path = ""
    output_path = ""
    template_dir = ""
    populate_excel(input_path,output_path,template_dir,0)
