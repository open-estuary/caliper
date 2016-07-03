#!/usr/bin/python
import os
import yaml
import openpyxl
import glob
path = "./Report_input"
del_list = ["peripheral"]

def populate_excel_to_yaml(row,c,dic,scenario_col,testcase_col,key,wb,file_name):
    for i in range(0,len(dic['results'][key].keys())):
        sheet_name = dic['results'][key].keys()[i]

        if sheet_name in del_list:
            continue
        sheet = wb.get_sheet_by_name(sheet_name)
        r = row
        head_row = row-1
        sheet.cell(row = head_row, column = c).value = file_name

        while(sheet.cell(row = r,column = scenario_col).value != None):

            try:
                if(dic['results'][key][sheet_name][sheet.cell(row = r,column = scenario_col).value][sheet.cell(row = r,column = testcase_col).value] >0):
                    sheet.cell(row = r,column = c).value = dic['results'][key][sheet_name][sheet.cell(row = r,column = scenario_col).value][sheet.cell(row = r,column = testcase_col).value]
                else:
                    sheet.cell(row=r, column=c).value = "Missing"
            except:
                sheet.cell(row = r,column = c).value = "Missing"
            r += 1


		

if __name__ == "__main__":
    keys = ["Performance","Functional"]
    col_start = 6
    row_start = 3
    scenario_column = 3
    testcase_column = 4
    col = col_start
    for key in keys:
        excel_name = key + ".xlsx"
        excel_name_save = key + "-Tests.xlsx"
        wb = openpyxl.load_workbook(excel_name)
        for files in glob.glob(os.path.join(path,"*yaml")):
            dic = yaml.load(open(files))
            file_name = files.split("/")[-1].split(".")[0]
            print(file_name)
            populate_excel_to_yaml(row_start,col,dic,scenario_column,testcase_column,key,wb,file_name)
            col += 1
        col = col_start
        wb.save(excel_name_save)
