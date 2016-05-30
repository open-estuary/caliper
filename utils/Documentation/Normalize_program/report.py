import openpyxl
import copy
import sys
import yaml
import dictionary
import pdb
import glob
import os
import json
from normalize import *
from report import *
from helper import *

def generateReport(excelName, reportName, function, heading_list):
    #copying the practical values to the ideal template yaml
    #dic_ideal = {}
    #copydic(function, dic_ideal, dic_practical)

    #open the excel sheet
    #try:
    #    wb = openpyxl.load_workbook(reportName+".xlsx")
    #except IOError:
    wb = openpyxl.Workbook()

    #checking if the sheet exits
    if reportName in wb.get_sheet_names():
        wb.remove_sheet(wb.get_sheet_by_name(reportName))

    #creating new sheet
    wb.create_sheet(index=0, title= reportName)
    sheet = wb.get_sheet_by_name(reportName)
    
    #initalise the row  
    r = 2
    c = 1
    SL = 1
    pos = 0
    size = len(heading_list)
    # initalise the column value
    prev_row = [2] * len(heading_list)
    present_row = [2] * len(heading_list)
    index = [1] * len(heading_list)
    index[size - 2] = 0

    #INSERTING THE HEADING LIST
    for heading in heading_list:
        pos += 1
        sheet.cell(row=1, column=pos).value = str(heading)

    #Creating the template yaml
    for TestName,functionName in function.iteritems():
        ideal = {}
        functionName(ideal)  
        c = c + 1
        prev_row[0] = present_row[size - 2]
        for top,category in ideal.iteritems():
            prev_row[1] = present_row[size - 2]
            if size > 2:
                c = c + 1
                for category,sub_category in category.iteritems():
                    prev_row[2] = present_row[size - 2]
                    if size > 3:
                        c = c + 1
                        for sub_category,scenario in sub_category.iteritems():
                            prev_row[3] = present_row[size - 2]
                            if size > 4:
                                c = c + 1
                                for scenario,key in scenario.iteritems():
                                    prev_row[4] = present_row[size - 2]
                                    if size > 5:
                                        c = c + 1
                                        for key,value in key.iteritems():
                                            prev_row[5] = present_row[size - 2]
                                            populate_row(sheet, prev_row[5], present_row[size - 2] - index[4], c, str(key))
                                            present_row[4] += 1
                                        c = c - 1
                                    populate_row(sheet, prev_row[4], present_row[size - 2] - index[3], c, str(scenario))
                                    present_row[3] += 1
                                c = c - 1
                            populate_row(sheet, prev_row[3], present_row[size - 2] - index[2] , c, str(sub_category))
                            present_row[2] += 1
                        c = c - 1
                    populate_row(sheet, prev_row[2], present_row[size - 2] - index[1], c, str(category))
                    present_row[1] += 1
                c = c -1
        c = c - 1

        for row_in in range(prev_row[0], present_row[size -2]):
            sheet.cell(row=row_in, column=c).value = str(SL)
            c = c + 1
            sheet.cell(row=row_in, column=c).value = str(TestName.strip('[]'))
            c = c - 1
        SL += 1
    wb.save(excelName)
    return

def populate_row(sheet, row_start, row_stop, current_column, value):
    for row_in in range(row_start, row_stop + 1) :
        sheet.cell(row=row_in, column=current_column).value = str(value)

def updateReport(excelName, reportName, platform, pos, size, function, copy, dic_practical):
    #copying the practical values to the ideal template yaml
    dic_ideal = {}

    copy(function, dic_ideal, dic_practical)
    wb = openpyxl.load_workbook(excelName)
    sheet = wb.get_sheet_by_name(reportName.split(".")[0])
    sheet.cell(row=1, column=pos).value = platform.split(".")[0]
    r = 2
    size = size - 1

    # initalise the column value
    prev_row = [2] * (pos - 1)
    present_row = [2] * (pos - 1)
    index = [1] * (pos -1)
    index[size - 2] = 0

    #Creating the template yaml
    for TestName,functionName in function.iteritems():
        ideal = {}
        functionName(ideal)  
        prev_row[0] = present_row[size - 2]
        for top,category in ideal.iteritems():
            prev_row[1] = present_row[size - 2]
            if size > 2:
                for category,sub_category in category.iteritems():
                    prev_row[2] = present_row[size - 2]
                    if size > 3:
                        for sub_category,scenario in sub_category.iteritems():
                            prev_row[3] = present_row[size - 2]
                            if size > 4:
                                for scenario,key in scenario.iteritems():
                                    prev_row[4] = present_row[size - 2]
                                    if size > 5:
                                        for key,value in key.iteritems():
                                            prev_row[5] = present_row[size - 2]
                                            populate_value(sheet, dic_ideal[top][category][sub_category][scenario][key], present_row[4], pos)
                                            present_row[4] += 1
                                        continue
                                    populate_value(sheet, dic_ideal[top][category][sub_category][scenario],present_row[3], pos)
                                    present_row[3] += 1
                                continue
                            populate_value(sheet, dic_ideal[top][category][sub_category], present_row[2] , pos)
                            present_row[2] += 1
                        continue
                    populate_value(sheet, dic_ideal[top][category], present_row[1] , pos)
                    present_row[1] += 1
                continue
    wb.save(excelName)
    return

def populate_value(sheet, ideal_value,row_in, col):
    if ideal_value > 0:
        sheet.cell(row=row_in, column=col).value = ideal_value
    else:
        # checking function is done here
        sheet.cell(row=row_in, column=col).value = "Missing"
'''
    for TestName,functionName in function.iteritems():
        c = pos
        ideal = {}
        functionName(ideal)
        for top,category in ideal.iteritems():                                   
            if top == 'results':                                                         
                for category,sub_category in category.iteritems():                       
                    for sub_category,scenario in sub_category.iteritems():               
                        for scenario,key in scenario.iteritems():                        
                            for key,value in key.iteritems():                            
                                if dic_ideal[top][category][sub_category][scenario][key] > 0:
                                    sheet.cell(row=r, column=c).value = dic_ideal[top][category][sub_category][scenario][key]
                                else:
                                    #checking function is done here
                                    sheet.cell(row=r, column=c).value = "Missing" 
                                r += 1
'''


