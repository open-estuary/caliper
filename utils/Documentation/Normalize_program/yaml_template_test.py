import openpyxl
import sys
import yaml
import dictionary
import glob
import os
import json
#path="/home/dikshit/yaml_files1/"
path="/home/dikshit/auto_verify/yaml_files"

def copydic(function, dic_ideal, dic_practical):
    #generate ideal template
    for key,value in function.iteritems():
        value(dic_ideal)
    for top,category in dic_practical.iteritems():
        if top == 'results':
            for category,sub_category in category.iteritems():
                for sub_category,scenario in sub_category.iteritems():
                    for scenario,key in scenario.iteritems():
                        for key,value in key.iteritems():
                             dic_ideal[top][category][sub_category][scenario][key] = value
    return

def generateReport(reportName, function, dic_practical):
    #copying the practical values to the ideal template yaml
    dic_ideal = {}
    copydic(function, dic_ideal, dic_practical)

    #open the excel sheet
    #try:
    #    wb = openpyxl.load_workbook(reportName+".xlsx")
	#except IOError:
    wb = openpyxl.Workbook()

    #checking if the sheet exits
    if reportName in wb.get_sheet_names():
        wb.remove_sheet(wb.get_sheet_by_name(reportName))
	
    #creating new sheet
    wb.create_sheet(index=0, title= reportName)
    sheet = wb.get_sheet_by_name(reportName)
    
    #initalise the row  
    r = 2
    c = 1
    SL = 1
    
    #Creating the template yaml
    sheet.cell(row=1, column=1).value = "SL NO"
    sheet.cell(row=1, column=2).value = "TOOLS"
    sheet.cell(row=1, column=3).value = "Scenario"
    sheet.cell(row=1, column=4).value = "Testcase"


    #Creating the template yaml
    for TestName,functionName in function.iteritems():
        print TestName
        ideal = {}
        functionName(ideal) 

	    #initalise the column value 
        prev_rc1 = r
        sheet.cell(row=r, column=c).value = str(SL)
        SL += 1
        c += 1
        sheet.cell(row=r, column=c).value = str(TestName)  
        c += 1
        for top,category in ideal.iteritems():                                   
            if top == 'results':                                                         
                for category,sub_category in category.iteritems():                       
                    for sub_category,scenario in sub_category.iteritems():               
                        for scenario,key in scenario.iteritems():                        
                                sheet.cell(row=r, column=c).value =  str(category) + "\n" + str(sub_category) + "\n" + str(scenario) 
                                c += 1
                                prev_rc2 = r
                                for key,value in key.iteritems():                            
                                        sheet.cell(row=r, column=c).value =  str(key) 
                                        r += 1
                                c -= 1
                                sheet.merge_cells(start_row=prev_rc2,start_column=c,end_row= r-1,end_column=c)
                                currentCell = sheet.cell(row=prev_rc2,column=c).coordinate
                                sheet[currentCell].alignment = openpyxl.styles.Alignment(horizontal='center')
                                sheet[currentCell].alignment = openpyxl.styles.Alignment(vertical='center')
   	    c -= 1
   	    sheet.merge_cells(start_row=prev_rc1,start_column=c,end_row=r-1,end_column=c)
   	    currentCell = sheet.cell(row=prev_rc1,column=c).coordinate
   	    sheet[currentCell].alignment = openpyxl.styles.Alignment(horizontal='center')
   	    sheet[currentCell].alignment = openpyxl.styles.Alignment(vertical='center')
   	    c -= 1
   	    sheet.merge_cells(start_row=prev_rc1,start_column=c,end_row=r-1,end_column=c)
   	    currentCell = sheet.cell(row=prev_rc1,column=c).coordinate
   	    sheet[currentCell].alignment = openpyxl.styles.Alignment(horizontal='center')
   	    sheet[currentCell].alignment = openpyxl.styles.Alignment(vertical='center')
   	wb.save(reportName+".xlsx")
    return

def updateReport(reportName, pos, function, dic_practical):
    #copying the practical values to the ideal template yaml
    dic_ideal = {}
    copydic(function, dic_ideal, dic_practical)
    wb = openpyxl.load_workbook("Consolidated.xlsx")
    sheet = wb.get_sheet_by_name("Consolidated")
    sheet.cell(row=1, column=pos).value = reportName.split(".")[0]
    r = 2
    for TestName,functionName in function.iteritems():
        c = pos
        ideal = {}
        functionName(ideal) 
        for top,category in ideal.iteritems():                                   
            if top == 'results':                                                         
                for category,sub_category in category.iteritems():                       
                    for sub_category,scenario in sub_category.iteritems():               
                        for scenario,key in scenario.iteritems():                        
                            for key,value in key.iteritems():                            
                                if dic_ideal[top][category][sub_category][scenario][key] > 0:
                                    sheet.cell(row=r, column=c).value = dic_ideal[top][category][sub_category][scenario][key]
                                else:
                                    #checking function is done here
                                    sheet.cell(row=r, column=c).value = "Missing" 
                                r += 1
    wb.save("Consolidated.xlsx")
    return

def normalize(function, dicList):
    #generate ideal template
    dic_ideal = {}
    for key,value in function.iteritems():
        value(dic_ideal)
    newDic = dictionary.convertDic(dic_ideal)
    for top,category in newDic.iteritems():
        if top == 'results':
            for category,sub_category in category.iteritems():
                    for sub_category,scenario in sub_category.iteritems():
                        dic = {}
                        value_Sub = [] 
                        value_Sub_Normalize = []
                        index = 0
                        for dic in dicList:
                            try:
                                value_Sub.append(dic[top][category][sub_category]['Total_Scores'])
                            except KeyError:
                                value_Sub.append(0)
                        for value in value_Sub:
                            if max(value_Sub) > 0:
                                if value > 0:
                                    value_Sub_Normalize.append(round(100*value/max(value_Sub),2))
                                else:
                                    value_Sub_Normalize.append(0)
                            else:
                                value_Sub_Normalize.append(0)

                        for dic in dicList:
                            try:
                                dic[top][category][sub_category]['Total_Scores'] = value_Sub_Normalize[index]
                                index+=1
                            except KeyError:
                                pass 
                        for scenario,division in scenario.iteritems():
                            if scenario != "Total_Scores":
                                for division,key in division.iteritems():
                                    if division == "Total_Scores":
                                        dic = {}
                                        value_Sub_Normalize = []
                                        value_Sub = [] 
                                        index = 0
                                        for dic in dicList:
                                            try:
                                                value_Sub.append(dic[top][category][sub_category][scenario][division])
                                            except KeyError:
                                                value_Sub.append(0)
                                        for value in value_Sub:
                                            if max(value_Sub) > 0:
                                                if value > 0:
                                                    value_Sub_Normalize.append(round(100*value/max(value_Sub),2))
                                                else:
                                                    value_Sub_Normalize.append(0)
                                            else:
                                                value_Sub_Normalize.append(0)

                                        for dic in dicList:
                                            try:
                                                dic[top][category][sub_category][scenario][division] = value_Sub_Normalize[index]
                                                index+=1
                                            except KeyError:
                                                pass
                                    else:
                                        for key,value in key.iteritems():
                                            dic = {}
                                            value_Sub_Normalize = []
                                            value_Sub = [] 
                                            index = 0
                                            for dic in dicList:
                                                try:
                                                    value_Sub.append(dic[top][category][sub_category][scenario][division][key])
                                                except KeyError:
                                                    value_Sub.append(0)
                                            for value in value_Sub:
                                                if max(value_Sub) > 0:
                                                    if value > 0:
                                                        value_Sub_Normalize.append(round(100*value/max(value_Sub),2))
                                                    else:
                                                        value_Sub_Normalize.append(0)
                                                else:
                                                    value_Sub_Normalize.append(0)

                                            for dic in dicList:
                                                try:
                                                    dic[top][category][sub_category][scenario][division][key] = value_Sub_Normalize[index]
                                                    index+=1
                                                except KeyError:
                                                   pass 
    for dic in dicList:
        del dic['results']['Functional']['peripheral']


def save(dicList):
    for dic in dicList:
        outputyaml = open("./Normalized_Files/"+dic['name']+"_score_post.yaml",'w')
        outputjson = open("./Normalized_Files/"+dic['name']+"_score_post.json",'w')
        outputyaml.write(yaml.dump(dic, default_flow_style=False))
        outputjson.write(json.dumps(dic))
        outputyaml.close()
        outputjson.close()


dic = {}
testfunction = {
    '[cachebench]' : dictionary.cachebench,
    '[compile]' : dictionary.complie,
    '[coremark]' : dictionary.coremark,
    '[dhrystone]' : dictionary.dhrystone,
    '[ebizzy]' : dictionary.ebizzy,
    '[fio]' : dictionary.fio,
    '[hadoop]' : dictionary.hadoop, 
    '[iozone]' : dictionary.iozone,
    '[iperf]' : dictionary.iperf,
    '[kselftest]' : dictionary.kselftest,
    '[linpack]' : dictionary.linpack,
    '[lmbench]' : dictionary.lmbench, 
    '[ltp]' : dictionary.ltp,
    '[memtester]' : dictionary.memtester,
    '[nbench]' : dictionary.nbench,
    '[netperf]' : dictionary.netperf, 
    '[openssl]' : dictionary.openssl,
    '[perf]' : dictionary.perf,
    '[rttest]' : dictionary.rttest,
    '[tinymembench]' : dictionary.tinymembench,
    '[scimark]' : dictionary.scimark,
    '[scimarkJava]' : dictionary.scimarkJava,
    '[sysbench]' : dictionary.sysbench,
    '[unzip]' : dictionary.unzip
}


#open the test list file  
testlist = open("testList","r")

#delete the testlist from the dictionary 
for test in testlist:
    if test[0] == '#': 
       del testfunction[test[1:].strip()]

#for key,value in testfunction.iteritems():
#    value(dic_ideal)                      
#filelist=[]                               
#pos = 5                                   
#generateReport("Consolidated", testfunction, dic_ideal)
#for filename in glob.glob(os.path.join(path, '*.yaml')):
#        #filelist.append(filename.split("/")[-1])           
#            filelist.append(filename)                           
#filelist.sort()                                         
dic_ideal = {}                            
for key,value in testfunction.iteritems():
    value(dic_ideal)               


if sys.argv[1] == "-n":
    dicList = []
    for filename in glob.glob(os.path.join(path, '*.yaml')):
        dicList.append(yaml.load(open(filename)))
    normalize(testfunction,dicList)
    save(dicList)

if sys.argv[1] == "-r":
    pos = 5
    generateReport("Consolidated", testfunction, dic_ideal)
    for filename in glob.glob(os.path.join(path, '*.yaml')):
        dic_practical = yaml.load(open(filename))
        updateReport(filename.split("/")[-1],pos,testfunction,dic_practical)
        pos+=1
